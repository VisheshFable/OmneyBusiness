import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sanction Screening Test Cases"

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Headers
headers = ["TC ID", "Test Case Description", "Module", "Positive/Negative", "Priority", "Test Data", "Expected Result"]

# Write headers
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Test Cases Data
test_cases = [
    # ========== SANCTION SCREENING API INTEGRATION ==========
    ("TC_SS_001", "Verify sanction screening API is called when admin submits Online channel transaction for disbursement", "Sanction Screening API", "Positive", "Critical", "Transaction booked via Online channel, Admin user credentials", "System calls Sanction Screening API and receives response successfully"),
    ("TC_SS_002", "Verify sanction screening API is called when admin submits App channel transaction for disbursement", "Sanction Screening API", "Positive", "Critical", "Transaction booked via Mobile App channel, Admin user credentials", "System calls Sanction Screening API and receives response successfully"),
    ("TC_SS_003", "Verify API request URL is correctly formatted with remitter name, nationality, DOB and other parameters", "Sanction Screening API", "Positive", "Critical", "Remitter: RITESH VIJAY GOSAVI, Nationality: RUSS, DOB: 02032002", "API URL: http://192.168.200.166/snc/index.php/api/remitFastScreening_json/RITESH~VIJAY~GOSAVI/RUSS/02032002/..."),
    ("TC_SS_004", "Verify transaction with 0% match is released for Last Mile processing", "Sanction Screening API", "Positive", "Critical", "Remitter name with no sanctions match, Expected match: 0%", "Transaction is released and sent to Last Mile for further processing"),
    ("TC_SS_005", "Verify transaction with 50% match is released for Last Mile processing", "Sanction Screening API", "Positive", "High", "Remitter with partial name match, Expected match: 50%", "Transaction is released for processing as match is below 95%"),
    ("TC_SS_006", "Verify transaction with 94% match is released for Last Mile processing", "Sanction Screening API", "Positive", "High", "Remitter with high partial match, Expected match: 94%", "Transaction is released for processing (boundary value - just below threshold)"),
    ("TC_SS_007", "Verify transaction with exactly 95% match is parked in Exception Queue", "Sanction Screening API", "Positive", "Critical", "Remitter name: RITESH KUMAR, Expected match: 95%", "Transaction is parked in Exception Queue with default 'False Positive' category"),
    ("TC_SS_008", "Verify transaction with 99% match is parked in Exception Queue", "Sanction Screening API", "Positive", "Critical", "Remitter name matching sanctions list, Expected match: 99%", "Transaction is parked in Exception Queue with default 'False Positive' category"),
    ("TC_SS_009", "Verify transaction with 100% match is parked in Exception Queue", "Sanction Screening API", "Positive", "Critical", "Remitter name exactly matching sanctions list, Expected match: 100%", "Transaction is parked in Exception Queue with default 'False Positive' category"),
    ("TC_SS_010", "Verify multiple Online/App transactions can be screened in batch when admin submits", "Sanction Screening API", "Positive", "High", "5 transactions from Online channel, 3 transactions from App channel", "All 8 transactions are screened individually and results processed correctly"),
    ("TC_SS_011", "Verify API response fields are correctly parsed (sanctionType, sanctionRecId, sanctionName, nationality, dob, passport, listedBy, matched)", "Sanction Screening API", "Positive", "High", "API response with all fields populated", "All response fields are correctly parsed and stored in system"),
    ("TC_SS_012", "Verify system handles API timeout gracefully", "Sanction Screening API", "Negative", "Critical", "Simulate API timeout (>30 seconds)", "System displays appropriate error message and transaction remains in pending state"),
    ("TC_SS_013", "Verify system handles API connection failure", "Sanction Screening API", "Negative", "Critical", "Simulate network failure to API endpoint", "System displays connection error, transaction not processed, logged for retry"),
    ("TC_SS_014", "Verify system handles invalid API response format", "Sanction Screening API", "Negative", "High", "Malformed JSON response from API", "System handles error gracefully, logs error, transaction remains pending"),
    ("TC_SS_015", "Verify system handles empty API response", "Sanction Screening API", "Negative", "High", "API returns empty data array: {\"data\": [], \"errorMessage\": \"\"}", "Transaction treated as 0% match and released for processing"),
    ("TC_SS_016", "Verify system handles API error message response", "Sanction Screening API", "Negative", "High", "API response: {\"data\": [], \"errorMessage\": \"Server Error\"}", "System logs error, displays message to admin, transaction not processed"),
    ("TC_SS_017", "Verify screening is not triggered for Branch transactions", "Sanction Screening API", "Positive", "High", "Transaction booked via Branch channel", "Screening logic does not trigger; uses existing Branch module screening"),
    ("TC_SS_018", "Verify special characters in remitter name are handled in API URL", "Sanction Screening API", "Negative", "Medium", "Remitter name: O'BRIEN MCDONALD-SMITH", "Special characters are URL encoded properly and API called successfully"),
    ("TC_SS_019", "Verify very long remitter name is handled", "Sanction Screening API", "Negative", "Medium", "Remitter name with 100+ characters", "API handles long name, response received successfully"),
    ("TC_SS_020", "Verify screening for remitter with multiple nationalities", "Sanction Screening API", "Positive", "Medium", "Remitter with dual nationality", "API correctly screens against all nationalities"),

    # ========== EXCEPTION QUEUE MASTER TOOL ==========
    ("TC_EQ_001", "Verify Exception Queue Master Tool is accessible by Compliance Manager", "Exception Queue Tool", "Positive", "Critical", "Compliance Manager user credentials", "Tool opens successfully with all features visible"),
    ("TC_EQ_002", "Verify Date/Calendar picker filter is mandatory", "Exception Queue Tool", "Positive", "Critical", "No date selected", "System prevents loading data until date is selected; validation message displayed"),
    ("TC_EQ_003", "Verify transactions are displayed when valid date is selected", "Exception Queue Tool", "Positive", "Critical", "Date with existing exception transactions", "All transactions for selected date are displayed in the queue"),
    ("TC_EQ_004", "Verify Status filter works for 'Hold' status", "Exception Queue Tool", "Positive", "High", "Date filter + Status: Hold", "Only transactions with Hold status are displayed"),
    ("TC_EQ_005", "Verify Status filter works for 'Reject' status", "Exception Queue Tool", "Positive", "High", "Date filter + Status: Reject", "Only transactions with Reject status are displayed"),
    ("TC_EQ_006", "Verify Status filter works for 'Release' status", "Exception Queue Tool", "Positive", "High", "Date filter + Status: Release", "Only transactions with Release status are displayed"),
    ("TC_EQ_007", "Verify all transactions displayed when no status filter applied", "Exception Queue Tool", "Positive", "High", "Date filter only, Status: None", "All transactions regardless of status are displayed"),
    ("TC_EQ_008", "Verify Excel download functionality", "Exception Queue Tool", "Positive", "High", "Click Download button with date filter applied", "Excel file downloads with all filtered transaction data"),
    ("TC_EQ_009", "Verify Excel contains all required columns", "Exception Queue Tool", "Positive", "High", "Downloaded Excel file", "Excel contains: Sr.No, Transaction Reference Number, Remitter Name, Beneficiary Name, Max Match (%), Bank Name, Branch Name, Account Number, Amount SGD"),
    ("TC_EQ_010", "Verify default category is 'False Positive' for new exception transactions", "Exception Queue Tool", "Positive", "Critical", "New transaction with 95%+ match", "Transaction displays with 'False Positive' category by default"),
    ("TC_EQ_011", "Verify category can be updated from 'False Positive' to 'True Positive'", "Exception Queue Tool", "Positive", "Critical", "Transaction with False Positive category", "Category updated to True Positive successfully and saved"),
    ("TC_EQ_012", "Verify category can be updated from 'True Positive' to 'False Positive'", "Exception Queue Tool", "Positive", "High", "Transaction with True Positive category", "Category updated to False Positive successfully and saved"),
    ("TC_EQ_013", "Verify category can be changed multiple times without limit", "Exception Queue Tool", "Positive", "Medium", "Change category 5 times on same transaction", "All category changes saved successfully without restriction"),
    ("TC_EQ_014", "Verify AML Details hyperlink opens popup with correct information", "Exception Queue Tool", "Positive", "Critical", "Transaction with AML details, Click 'view' hyperlink", "Popup displays: Sanction Name, Remitter Name, Beneficiary Name, Matched %, Bene Remit, Listed By, Nationality, Remitter Country, DOB, Passport, Sanction Rec ID, Sanction Type, Value Date, Bene Account, Bank Name, Branch Name, Amount, Payment Mode, Status, Remarks"),
    ("TC_EQ_015", "Verify 'Bene Remit' field shows 'Remitter' when remitter name matched", "Exception Queue Tool", "Positive", "High", "Transaction where remitter name triggered match", "Bene Remit field displays 'Remitter'"),
    ("TC_EQ_016", "Verify 'Bene Remit' field shows 'Beneficiary' when beneficiary name matched", "Exception Queue Tool", "Positive", "High", "Transaction where beneficiary name triggered match", "Bene Remit field displays 'Beneficiary'"),
    ("TC_EQ_017", "Verify Authorize action sends transaction to Last Mile for disbursement", "Exception Queue Tool", "Positive", "Critical", "Select transaction, Click Authorize", "Transaction status updated, sent to Last Mile for processing"),
    ("TC_EQ_018", "Verify multiple transactions can be authorized at once", "Exception Queue Tool", "Positive", "High", "Select 5 transactions, Click Authorize", "All 5 transactions authorized and sent to Last Mile"),
    ("TC_EQ_019", "Verify Authorize reason is non-mandatory", "Exception Queue Tool", "Positive", "Medium", "Authorize without entering reason", "Transaction authorized successfully without reason"),
    ("TC_EQ_020", "Verify Reject action cancels transaction and treats as refund", "Exception Queue Tool", "Positive", "Critical", "Select transaction, Enter reason, Click Reject", "Transaction cancelled, marked for refund processing"),
    ("TC_EQ_021", "Verify Reject reason is mandatory", "Exception Queue Tool", "Negative", "Critical", "Try to reject without entering reason", "System prevents rejection, displays 'Reason is mandatory' message"),
    ("TC_EQ_022", "Verify Hold action holds transaction for later review", "Exception Queue Tool", "Positive", "Critical", "Select transaction, Enter reason, Click Hold", "Transaction status updated to Hold, available for later action"),
    ("TC_EQ_023", "Verify Hold reason is mandatory", "Exception Queue Tool", "Negative", "Critical", "Try to hold without entering reason", "System prevents hold action, displays 'Reason is mandatory' message"),
    ("TC_EQ_024", "Verify Hold transaction can be later Authorized", "Exception Queue Tool", "Positive", "High", "Transaction on Hold, Click Authorize", "Transaction changes from Hold to Authorized, sent to Last Mile"),
    ("TC_EQ_025", "Verify Hold transaction can be later Rejected", "Exception Queue Tool", "Positive", "High", "Transaction on Hold, Enter reason, Click Reject", "Transaction changes from Hold to Rejected, processed as refund"),
    ("TC_EQ_026", "Verify Hold filter displays only held transactions", "Exception Queue Tool", "Positive", "High", "Apply Hold filter", "Only transactions with Hold status are displayed"),
    ("TC_EQ_027", "Verify comments field is mandatory for all status updates", "Exception Queue Tool", "Positive", "High", "Update status without comments", "System prevents update, displays mandatory comment message"),
    ("TC_EQ_028", "Verify SUBMIT button updates transaction status", "Exception Queue Tool", "Positive", "Critical", "Select transaction, change status, Click Submit", "Transaction status updated and saved"),
    ("TC_EQ_029", "Verify VIEW button displays transaction details", "Exception Queue Tool", "Positive", "High", "Select transaction, Click View", "Transaction details popup/page displayed"),
    ("TC_EQ_030", "Verify DOWNLOAD button exports data based on date filter", "Exception Queue Tool", "Positive", "High", "Apply date filter, Click Download", "Excel file downloads with data for selected date range"),
    ("TC_EQ_031", "Verify Transaction Reference Number hyperlink downloads PDF", "Exception Queue Tool", "Positive", "High", "Click on hyperlinked Transaction Reference Number", "Transaction detail PDF downloads to computer"),
    ("TC_EQ_032", "Verify single maker user concept - only one user can edit at a time", "Exception Queue Tool", "Positive", "Critical", "Two users try to edit same transaction", "Only first user allowed; second user sees lock message"),
    ("TC_EQ_033", "Verify no data message when no exceptions for selected date", "Exception Queue Tool", "Positive", "Medium", "Date with no exception transactions", "Message 'No records found' displayed"),
    ("TC_EQ_034", "Verify future date selection is handled", "Exception Queue Tool", "Negative", "Low", "Select future date", "No records found or date validation prevents future date"),
    ("TC_EQ_035", "Verify date range selection works correctly", "Exception Queue Tool", "Positive", "Medium", "Select date range: 01/01/2025 to 15/01/2025", "All transactions within date range displayed"),
    ("TC_EQ_036", "Verify category change is captured in corresponding report", "Exception Queue Tool", "Positive", "High", "Change category from False Positive to True Positive", "Transaction appears in True Positive report after category change"),
    ("TC_EQ_037", "Verify unauthorized user cannot access Exception Queue Tool", "Exception Queue Tool", "Negative", "Critical", "Non-compliance manager credentials", "Access denied message displayed"),
    ("TC_EQ_038", "Verify session timeout handling", "Exception Queue Tool", "Negative", "Medium", "Leave tool idle for extended period", "Session expires, user redirected to login"),
    ("TC_EQ_039", "Verify empty comments field validation", "Exception Queue Tool", "Negative", "High", "Enter only spaces in comments field", "Validation error - meaningful comment required"),
    ("TC_EQ_040", "Verify maximum character limit for comments", "Exception Queue Tool", "Negative", "Low", "Enter 5000+ characters in comments", "System truncates or shows max character limit message"),

    # ========== RE-PUSH MECHANISM ==========
    ("TC_RP_001", "Verify system scheduler checks for stuck transactions after authorization", "Re-push Mechanism", "Positive", "Critical", "Authorized transaction stuck due to API unavailability", "Scheduler identifies stuck transaction and retries push to Last Mile"),
    ("TC_RP_002", "Verify pending transactions are automatically pushed when API becomes available", "Re-push Mechanism", "Positive", "Critical", "Transactions stuck when sanction server was busy", "Transactions automatically pushed to Last Mile when server available"),
    ("TC_RP_003", "Verify scheduler runs at configured interval", "Re-push Mechanism", "Positive", "High", "Configure scheduler for 5-minute intervals", "Scheduler runs every 5 minutes to check pending transactions"),
    ("TC_RP_004", "Verify re-push does not duplicate transactions", "Re-push Mechanism", "Positive", "Critical", "Transaction already pushed to Last Mile", "No duplicate transaction created in Last Mile"),
    ("TC_RP_005", "Verify re-push logging for audit trail", "Re-push Mechanism", "Positive", "High", "Re-push mechanism executes", "All re-push attempts logged with timestamp and status"),
    ("TC_RP_006", "Verify maximum retry attempts for stuck transactions", "Re-push Mechanism", "Negative", "High", "Transaction fails re-push 10 times", "Transaction marked for manual intervention after max retries"),
    ("TC_RP_007", "Verify notification/alert for continuously failing re-push", "Re-push Mechanism", "Positive", "Medium", "Transaction fails multiple re-push attempts", "Alert sent to admin for manual review"),

    # ========== SANCTION SCREENING REPORT ==========
    ("TC_RPT_001", "Verify Sanction Screening Report dashboard displays correct counts", "Sanction Screening Report", "Positive", "Critical", "Date range with known screening data", "Dashboard shows correct counts for all categories: No Match, False Positive, True Positive for Remitter and Beneficiary"),
    ("TC_RPT_002", "Verify No Match - Remitter count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on No Match - Remitter count", "Excel downloads with No Match Remitter data"),
    ("TC_RPT_003", "Verify No Match - Beneficiary count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on No Match - Beneficiary count", "Excel downloads with No Match Beneficiary data"),
    ("TC_RPT_004", "Verify False Positive - Remitter count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on False Positive - Remitter count", "Excel downloads with False Positive Remitter data"),
    ("TC_RPT_005", "Verify False Positive - Beneficiary count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on False Positive - Beneficiary count", "Excel downloads with False Positive Beneficiary data"),
    ("TC_RPT_006", "Verify True Positive - Remitter count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on True Positive - Remitter count", "Excel downloads with True Positive Remitter data"),
    ("TC_RPT_007", "Verify True Positive - Beneficiary count is hyperlinked", "Sanction Screening Report", "Positive", "High", "Click on True Positive - Beneficiary count", "Excel downloads with True Positive Beneficiary data"),
    ("TC_RPT_008", "Verify Download PDF functionality", "Sanction Screening Report", "Positive", "High", "Click Download PDF button", "PDF report with summary data downloads"),
    ("TC_RPT_009", "Verify PDF report format matches sample wireframe", "Sanction Screening Report", "Positive", "Medium", "Downloaded PDF report", "PDF matches expected format with all categories and counts"),
    ("TC_RPT_010", "Verify report displays data for all modules (Online, App, Branch)", "Sanction Screening Report", "Positive", "High", "Date range with transactions from all modules", "Report includes screening data from all channels"),
    ("TC_RPT_011", "Verify zero counts displayed when no data for category", "Sanction Screening Report", "Positive", "Medium", "Date with no True Positive matches", "True Positive count shows 0"),
    ("TC_RPT_012", "Verify date filter works correctly", "Sanction Screening Report", "Positive", "High", "Apply date filter for specific range", "Only data within date range displayed"),

    # ========== FALSE POSITIVE REPORT ==========
    ("TC_FPR_001", "Verify False Positive Report displays correct data", "False Positive Report", "Positive", "Critical", "Date range with False Positive transactions", "Report shows all False Positive flagged transactions"),
    ("TC_FPR_002", "Verify Transaction Reference Number is hyperlinked in report", "False Positive Report", "Positive", "High", "View False Positive Report", "Transaction Reference Numbers are clickable hyperlinks"),
    ("TC_FPR_003", "Verify clicking Transaction Reference downloads PDF", "False Positive Report", "Positive", "High", "Click on Transaction Reference Number hyperlink", "Transaction detail PDF downloads to computer"),
    ("TC_FPR_004", "Verify separate Excel download for Remitter False Positives", "False Positive Report", "Positive", "High", "Download Remitter False Positive data", "Excel file contains only Remitter False Positive records"),
    ("TC_FPR_005", "Verify separate Excel download for Beneficiary False Positives", "False Positive Report", "Positive", "High", "Download Beneficiary False Positive data", "Excel file contains only Beneficiary False Positive records"),
    ("TC_FPR_006", "Verify Excel columns match wireframe specification", "False Positive Report", "Positive", "Medium", "Downloaded Excel file", "Excel contains: Sr.No, Transaction Reference Number, Remitter Name, Beneficiary Name, Max Match %, Bank Name, Branch Name, Account Number, Amount SGD"),
    ("TC_FPR_007", "Verify PDF download for False Positive Report", "False Positive Report", "Positive", "High", "Click PDF download", "PDF report downloads with correct formatting"),
    ("TC_FPR_008", "Verify empty report message when no False Positives", "False Positive Report", "Positive", "Medium", "Date with no False Positive transactions", "Message 'No records found' displayed"),
    ("TC_FPR_009", "Verify report updates when transaction category changes from False to True Positive", "False Positive Report", "Positive", "High", "Change transaction category to True Positive", "Transaction no longer appears in False Positive Report"),

    # ========== TRUE POSITIVE REPORT ==========
    ("TC_TPR_001", "Verify True Positive Report displays correct data", "True Positive Report", "Positive", "Critical", "Date range with True Positive transactions", "Report shows all True Positive flagged transactions"),
    ("TC_TPR_002", "Verify Transaction Reference Number is hyperlinked", "True Positive Report", "Positive", "High", "View True Positive Report", "Transaction Reference Numbers are clickable hyperlinks"),
    ("TC_TPR_003", "Verify clicking Transaction Reference downloads PDF", "True Positive Report", "Positive", "High", "Click on Transaction Reference Number hyperlink", "Transaction detail PDF downloads to computer"),
    ("TC_TPR_004", "Verify separate Excel download for Remitter True Positives", "True Positive Report", "Positive", "High", "Download Remitter True Positive data", "Excel file contains only Remitter True Positive records"),
    ("TC_TPR_005", "Verify separate Excel download for Beneficiary True Positives", "True Positive Report", "Positive", "High", "Download Beneficiary True Positive data", "Excel file contains only Beneficiary True Positive records"),
    ("TC_TPR_006", "Verify report format matches False Positive Report wireframe", "True Positive Report", "Positive", "Medium", "View True Positive Report", "Format matches False Positive Report layout"),
    ("TC_TPR_007", "Verify PDF download for True Positive Report", "True Positive Report", "Positive", "High", "Click PDF download", "PDF report downloads with correct formatting"),
    ("TC_TPR_008", "Verify empty report message when no True Positives", "True Positive Report", "Positive", "Medium", "Date with no True Positive transactions", "Message 'No records found' displayed"),

    # ========== NO MATCH REPORT ==========
    ("TC_NMR_001", "Verify No Match Report displays transactions with 0% match", "No Match Report", "Positive", "Critical", "Date range with 0% match transactions", "Report shows all transactions with 0% sanction match"),
    ("TC_NMR_002", "Verify Transaction Reference Number is hyperlinked", "No Match Report", "Positive", "High", "View No Match Report", "Transaction Reference Numbers are clickable hyperlinks"),
    ("TC_NMR_003", "Verify clicking Transaction Reference downloads PDF", "No Match Report", "Positive", "High", "Click on Transaction Reference Number hyperlink", "Transaction detail PDF downloads to computer"),
    ("TC_NMR_004", "Verify separate Excel download for Remitter No Match", "No Match Report", "Positive", "High", "Download Remitter No Match data", "Excel file contains only Remitter No Match records"),
    ("TC_NMR_005", "Verify separate Excel download for Beneficiary No Match", "No Match Report", "Positive", "High", "Download Beneficiary No Match data", "Excel file contains only Beneficiary No Match records"),
    ("TC_NMR_006", "Verify report format matches wireframe specification", "No Match Report", "Positive", "Medium", "View No Match Report", "Format matches expected layout"),
    ("TC_NMR_007", "Verify PDF download for No Match Report", "No Match Report", "Positive", "High", "Click PDF download", "PDF report downloads with correct formatting"),
    ("TC_NMR_008", "Verify empty report message when no No Match transactions", "No Match Report", "Positive", "Medium", "Date with no 0% match transactions", "Message 'No records found' displayed"),

    # ========== INTEGRATION TEST CASES ==========
    ("TC_INT_001", "Verify end-to-end flow: Online transaction booking to Last Mile processing (No Match)", "Integration", "Positive", "Critical", "Book Online transaction with unique name (0% match)", "Transaction screened, 0% match, released to Last Mile, processed successfully"),
    ("TC_INT_002", "Verify end-to-end flow: App transaction booking to Exception Queue (95%+ Match)", "Integration", "Positive", "Critical", "Book App transaction with sanctioned name (95%+ match)", "Transaction screened, parked in Exception Queue, Compliance review required"),
    ("TC_INT_003", "Verify end-to-end flow: Exception Queue to Last Mile after Authorization", "Integration", "Positive", "Critical", "Transaction in Exception Queue, Compliance authorizes", "Transaction moves from Exception Queue to Last Mile for disbursement"),
    ("TC_INT_004", "Verify end-to-end flow: Exception Queue to Refund after Rejection", "Integration", "Positive", "Critical", "Transaction in Exception Queue, Compliance rejects with reason", "Transaction cancelled, refund initiated to customer"),
    ("TC_INT_005", "Verify screening logic consistency between Online and Branch modules", "Integration", "Positive", "High", "Same remitter name in Online and Branch transactions", "Both transactions show same match percentage"),
    ("TC_INT_006", "Verify data consistency across Exception Queue and Reports", "Integration", "Positive", "High", "Transaction flagged as True Positive", "Transaction appears correctly in True Positive Report"),
    ("TC_INT_007", "Verify concurrent transaction processing", "Integration", "Positive", "High", "100 transactions submitted simultaneously", "All transactions screened correctly without data loss or corruption"),
    ("TC_INT_008", "Verify transaction audit trail from booking to final status", "Integration", "Positive", "High", "Complete transaction lifecycle", "All status changes logged with timestamp, user, and action"),

    # ========== SECURITY TEST CASES ==========
    ("TC_SEC_001", "Verify API authentication for Sanction Screening API", "Security", "Positive", "Critical", "Valid API credentials/token", "API authenticates successfully and returns response"),
    ("TC_SEC_002", "Verify API request fails with invalid credentials", "Security", "Negative", "Critical", "Invalid API credentials/token", "API returns authentication error, no data exposed"),
    ("TC_SEC_003", "Verify SQL injection prevention in search/filter fields", "Security", "Negative", "Critical", "Input: ' OR '1'='1 in filter fields", "No SQL injection executed, input sanitized"),
    ("TC_SEC_004", "Verify XSS prevention in comments and input fields", "Security", "Negative", "Critical", "Input: <script>alert('XSS')</script>", "Script not executed, input escaped/sanitized"),
    ("TC_SEC_005", "Verify role-based access control for Exception Queue", "Security", "Positive", "Critical", "Different user roles accessing tool", "Only authorized roles can access Exception Queue features"),
    ("TC_SEC_006", "Verify sensitive data encryption in transit", "Security", "Positive", "Critical", "Monitor API traffic", "All API calls use HTTPS, data encrypted"),
    ("TC_SEC_007", "Verify audit logging for all user actions", "Security", "Positive", "High", "Perform various actions in Exception Queue", "All actions logged with user ID, timestamp, and action details"),
    ("TC_SEC_008", "Verify session management and timeout", "Security", "Positive", "High", "Leave session idle beyond timeout", "Session expires, user must re-authenticate"),

    # ========== PERFORMANCE TEST CASES ==========
    ("TC_PERF_001", "Verify Sanction Screening API response time under normal load", "Performance", "Positive", "High", "Single API call", "API response received within 3 seconds"),
    ("TC_PERF_002", "Verify bulk transaction screening performance", "Performance", "Positive", "High", "Submit 100 transactions for screening", "All transactions screened within acceptable time (< 60 seconds)"),
    ("TC_PERF_003", "Verify Exception Queue tool loading time", "Performance", "Positive", "Medium", "Load Exception Queue with 1000 records", "Page loads within 5 seconds"),
    ("TC_PERF_004", "Verify report generation time for large datasets", "Performance", "Positive", "Medium", "Generate report for 1 month of data (10000+ records)", "Report generated within 30 seconds"),
    ("TC_PERF_005", "Verify Excel download performance for large datasets", "Performance", "Positive", "Medium", "Download Excel with 5000 records", "Excel file downloads within 15 seconds"),
    ("TC_PERF_006", "Verify system behavior under peak load", "Performance", "Positive", "High", "Simulate 50 concurrent users", "System remains responsive, no errors"),

    # ========== USABILITY TEST CASES ==========
    ("TC_USA_001", "Verify Exception Queue tool UI matches wireframe design", "Usability", "Positive", "Medium", "Compare UI with design mockups", "UI matches approved wireframe design"),
    ("TC_USA_002", "Verify error messages are user-friendly", "Usability", "Positive", "Medium", "Trigger various error conditions", "Error messages are clear and actionable"),
    ("TC_USA_003", "Verify mandatory field indicators are visible", "Usability", "Positive", "Medium", "View forms with mandatory fields", "Mandatory fields clearly marked with asterisk or indicator"),
    ("TC_USA_004", "Verify date picker functionality", "Usability", "Positive", "Low", "Use date picker to select dates", "Date picker is intuitive and functions correctly"),
    ("TC_USA_005", "Verify report download progress indicator", "Usability", "Positive", "Low", "Download large report", "Progress indicator shown during download"),
    ("TC_USA_006", "Verify browser compatibility (Chrome, Firefox, Edge)", "Usability", "Positive", "Medium", "Access tool from different browsers", "Tool functions correctly in all major browsers"),

    # ========== BOUNDARY VALUE ANALYSIS ==========
    ("TC_BVA_001", "Verify match percentage at lower boundary (0%)", "Boundary Analysis", "Positive", "High", "Remitter with 0% match", "Transaction released for processing"),
    ("TC_BVA_002", "Verify match percentage just below threshold (94%)", "Boundary Analysis", "Positive", "High", "Remitter with 94% match", "Transaction released for processing"),
    ("TC_BVA_003", "Verify match percentage at threshold (95%)", "Boundary Analysis", "Positive", "Critical", "Remitter with exactly 95% match", "Transaction parked in Exception Queue"),
    ("TC_BVA_004", "Verify match percentage just above threshold (96%)", "Boundary Analysis", "Positive", "High", "Remitter with 96% match", "Transaction parked in Exception Queue"),
    ("TC_BVA_005", "Verify match percentage at upper boundary (100%)", "Boundary Analysis", "Positive", "High", "Remitter with 100% match", "Transaction parked in Exception Queue"),
    ("TC_BVA_006", "Verify minimum transaction amount handling", "Boundary Analysis", "Positive", "Medium", "Transaction amount: 0.01 SGD", "Transaction screened normally"),
    ("TC_BVA_007", "Verify maximum transaction amount handling", "Boundary Analysis", "Positive", "Medium", "Transaction amount: 999999.99 SGD", "Transaction screened normally"),

    # ========== NEGATIVE/ERROR HANDLING ==========
    ("TC_NEG_001", "Verify handling of special characters in remitter name", "Error Handling", "Negative", "Medium", "Remitter name: @#$%^&*()", "System handles gracefully, no crash"),
    ("TC_NEG_002", "Verify handling of numeric-only remitter name", "Error Handling", "Negative", "Low", "Remitter name: 12345678", "System screens name, API handles appropriately"),
    ("TC_NEG_003", "Verify handling of empty remitter name", "Error Handling", "Negative", "High", "Remitter name: (empty/blank)", "Validation error or handled gracefully"),
    ("TC_NEG_004", "Verify handling of invalid date format in DOB", "Error Handling", "Negative", "Medium", "DOB: 99/99/9999", "Error handled, logged appropriately"),
    ("TC_NEG_005", "Verify handling of missing nationality", "Error Handling", "Negative", "Medium", "Nationality field empty", "System handles missing data gracefully"),
    ("TC_NEG_006", "Verify database connection failure handling", "Error Handling", "Negative", "Critical", "Simulate database unavailability", "User-friendly error message, transaction preserved"),
    ("TC_NEG_007", "Verify handling of concurrent status update on same transaction", "Error Handling", "Negative", "High", "Two users update same transaction simultaneously", "First update succeeds, second notified of conflict"),
    ("TC_NEG_008", "Verify handling of malformed Transaction Reference Number", "Error Handling", "Negative", "Medium", "Invalid format reference number", "System handles gracefully, logs error"),

    # ========== REGULATORY COMPLIANCE ==========
    ("TC_REG_001", "Verify OFAC sanctions list integration", "Compliance", "Positive", "Critical", "Name on OFAC list", "Match detected against OFAC list"),
    ("TC_REG_002", "Verify UN sanctions list integration", "Compliance", "Positive", "Critical", "Name on UN sanctions list", "Match detected against UN list"),
    ("TC_REG_003", "Verify EU sanctions list integration", "Compliance", "Positive", "Critical", "Name on EU sanctions list", "Match detected against EU list"),
    ("TC_REG_004", "Verify RBI sanctions list integration (if applicable)", "Compliance", "Positive", "Critical", "Name on RBI list", "Match detected against RBI list"),
    ("TC_REG_005", "Verify PEP (Politically Exposed Person) detection", "Compliance", "Positive", "Critical", "Name matching PEP in DOWJONES-PEPs", "PEP match detected, parked in Exception Queue"),
    ("TC_REG_006", "Verify audit trail meets regulatory requirements", "Compliance", "Positive", "High", "Review audit logs", "All required information logged for compliance audit"),
    ("TC_REG_007", "Verify data retention policy compliance", "Compliance", "Positive", "Medium", "Check historical data availability", "Data retained as per regulatory requirements"),
]

# Write test cases
for row, tc in enumerate(test_cases, 2):
    for col, value in enumerate(tc, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.border = thin_border
        cell.alignment = wrap_alignment

# Adjust column widths
column_widths = [12, 70, 25, 18, 12, 50, 60]
for i, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

# Set row height for header
ws.row_dimensions[1].height = 25

# Freeze first row
ws.freeze_panes = 'A2'

# Add conditional formatting for Priority
from openpyxl.formatting.rule import FormulaRule

# Color coding for priority (using PatternFill for data bars alternative)
critical_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
high_fill = PatternFill(start_color="FFB347", end_color="FFB347", fill_type="solid")
medium_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
low_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

# Apply colors to priority column
for row in range(2, len(test_cases) + 2):
    priority_cell = ws.cell(row=row, column=5)
    if priority_cell.value == "Critical":
        priority_cell.fill = critical_fill
    elif priority_cell.value == "High":
        priority_cell.fill = high_fill
    elif priority_cell.value == "Medium":
        priority_cell.fill = medium_fill
    elif priority_cell.value == "Low":
        priority_cell.fill = low_fill

# Save workbook
output_path = r"D:\Vishesh\OmneyBusiness\Sanction_Screening_Test_Cases.xlsx"
wb.save(output_path)
print(f"Test cases saved to: {output_path}")
print(f"Total test cases: {len(test_cases)}")

# Print summary
modules = {}
positive_count = 0
negative_count = 0
priorities = {"Critical": 0, "High": 0, "Medium": 0, "Low": 0}

for tc in test_cases:
    module = tc[2]
    pos_neg = tc[3]
    priority = tc[4]

    modules[module] = modules.get(module, 0) + 1
    if pos_neg == "Positive":
        positive_count += 1
    else:
        negative_count += 1
    priorities[priority] = priorities.get(priority, 0) + 1

print("\n=== TEST CASE SUMMARY ===")
print(f"\nTotal Test Cases: {len(test_cases)}")
print(f"Positive Cases: {positive_count}")
print(f"Negative Cases: {negative_count}")
print("\nBy Priority:")
for p, count in priorities.items():
    print(f"  {p}: {count}")
print("\nBy Module:")
for m, count in sorted(modules.items(), key=lambda x: -x[1]):
    print(f"  {m}: {count}")
